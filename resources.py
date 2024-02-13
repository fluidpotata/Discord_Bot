import openpyxl
import random
import sqlite3
# import os
# import mysql.connector

wb = openpyxl.load_workbook('master.xlsx')


def connectDB():
    """ Old database lost, adding a local db """
    # config = {
    #     'user': "sql12657836",
    #     'password': "XUJAZEqxsI",
    #     'host': "sql12.freemysqlhosting.net",
    #     'database': "sql12657836",
    #     'raise_on_warnings': True
    # }
    # cnx = mysql.connector.connect(**config)
    cnx = sqlite3.connect("db/bot.db")
    return cnx


async def materials(ctx, code: str):
    """
    Need to configure with database
    """
    ws = wb['resources']
    n = int(ws[1][0].value)
    for i in range(2, n + 1):
        if str(ws[i][0].value) == code.upper():
            l = int(ws[i][1].value)
            output = ""
            for j in range(2, l + 1):
                x = str(ws[i][j].value)
                if x != 'None':
                    output += str(ws[i][j].value)
                    output += "\n"
            return True, output
    return False, "Not available yet!"


def new_func(code):
    return code.upper()


async def facultySearch(name):
    connection = connectDB()
    cursor = connection.cursor()
    if len(name) == 3:
        cursor.execute(
            f"""select * from facinfos where initial like "%{name}%";""")
    if len(name) > 3:
        cursor.execute(
            f"""select * from facinfos where name like "%{name}%";""")
    result = cursor.fetchall()
    cursor.close()
    connection.close()
    return result


def mid(data):
    try:
        connection = connectDB()
        cursor = connection.cursor()
        cursor.execute(
            f"select * from final_f23 where course = '{data[0]}' and section = '{data[1]}';"
        )
        out = cursor.fetchall()
        out = out[0]
        result = f"""```Course: {out[0]}\nSection: {out[1]}\nDate: {out[2]}\nTime: {out[3]}-{out[4]}\nRoom: {out[5]}\nMode: {out[6]}```\n"""
        cursor.close()
        connection.close()
        return result
    except:
        return "Database busy! Try again after 5mins!"
